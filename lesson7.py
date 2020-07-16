# -*- coding: utf-8 -*-
# @Time ： 2020/7/15 21:00
# @Auth ： wannru
# @File ：lesson7.py
# @QQ ：943703484
# @Company：湖南省零檬信息技术有限公司

import requests
import openpyxl

#用来读取测试用例函数
def read_data(filename,sheetname):
    we = openpyxl.load_workbook(filename)     #加载工作薄 -- 写入文档名字
    sheet = we[sheetname]      #获取表单
    max_row = sheet.max_row    #获取做大行数
    case_list = []    #创建空列表，存放测试用例
    for i in range(2,max_row+1):
        dict1 = dict(
        case_id = sheet.cell(row=i,column=1).value,   #获取case_id
        url = sheet.cell(row=i,column=5).value,  #获取url
        data = sheet.cell(row=i,column=6).value,  #获取data
        expect = sheet.cell(row=i,column=7).value  #获取expect
        )
        case_list.append(dict1)  #每循环一次，就把读取到的字典数据存放到这个list
    return case_list   #返回测试用例列表

#执行接口函数
def api_fun(url,data):
    headers_log = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}  #请求体-字典
    res = requests.post(url=url,json=data,headers=headers_log)  #接收post方法的结果
    response = res.json()#响应正文
    return response

# 写入结果
def write_result(filename,sheetname,row,column,final_result):
    we = openpyxl.load_workbook(filename)
    sheet = we[sheetname]
    sheet.cell(row=row,column=column).value = final_result   #写入结果
    we.save('test_case_api.xlsx')  #保存
# write_result('test_case_api.xlsx','login',3,8,'Failed')

def execute_fun(filename,sheetname):
    cases = read_data(filename,sheetname)  #调用读取测试用例，获取所有测试用例数据保存到变量
    for case in cases:
        case_id = case.get('case_id')  #case['case_id']
        url = case.get('url')
        data =eval(case.get('data'))   #eval(） 运行被字符串包裹的表达式---去掉字符串的引号
        expect =eval(case.get('expect'))    #获取预期结果
        expect_msg = expect.get('msg')  #获取预期结果中的msg
        real_result = api_fun(url=url,data=data)  #调用发送接口请求函数，返回结果用变量real_result接收
        real_msg = real_result.get('msg')  #获取实际结果中的msg
        print('预期结果中的msg:{}'.format(expect_msg))
        print('实际结果中的msg:{}'.format(real_msg))
        if real_msg == expect_msg:
            print('第{}条测试用例执行通过！'.format(case_id))
            final_re = 'Passed'
        else:
            print('第{}条测试用例执行不通过！'.format(case_id))
            final_re = 'Failed'
        write_result(filename,sheetname,case_id+1,8,final_re)
        print('*'*25)
execute_fun('test_case_api.xlsx','register')
